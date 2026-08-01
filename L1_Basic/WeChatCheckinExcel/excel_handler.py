import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from config import EXCEL_FILE_PATH, EXCEL_COLUMNS

def ensure_excel_exists(file_path: str = EXCEL_FILE_PATH, columns: list = EXCEL_COLUMNS):
    """
    检查 Excel 文件是否存在，若不存在则创建并初始化表头格式。
    """
    dir_path = os.path.dirname(file_path)
    if dir_path and not os.path.exists(dir_path):
        os.makedirs(dir_path, exist_ok=True)

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "打卡预约记录"

        # 写入表头
        ws.append(columns)

        # 表头样式设置（灰色背景，粗体，居中）
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(columns)):
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

        # 设置列宽
        column_widths = {
            "A": 20, # 接收时间
            "B": 18, # 消息来源(微信名)
            "C": 20, # 账号
            "D": 20, # 密码
            "E": 20, # 预约时间
            "F": 12, # 有无绑定
            "G": 15, # 是否是diploma
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(file_path)
        print(f"[Excel] 成功初始化表格文件: {file_path}")

def append_appointment_record(sender: str, parsed_data: dict, file_path: str = EXCEL_FILE_PATH):
    """
    追加一条打卡预约数据到 Excel 表格中。
    """
    ensure_excel_exists(file_path)

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_values = [
        now_str,
        sender,
        parsed_data.get("account", ""),
        parsed_data.get("password", ""),
        parsed_data.get("time", ""),
        parsed_data.get("is_bound", ""),
        parsed_data.get("is_diploma", "")
    ]

    wb = load_workbook(file_path)
    ws = wb.active

    ws.append(row_values)

    # 给新增行设置对齐样式
    last_row = ws.max_row
    align_center = Alignment(horizontal="center", vertical="center")
    for cell in ws[last_row]:
        cell.alignment = align_center

    wb.save(file_path)
    print(f"[Excel] 已成功写入 1 条记录 (来源: {sender}, 账号: {parsed_data.get('account')})")
