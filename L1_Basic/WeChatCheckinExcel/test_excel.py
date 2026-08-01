import os
import unittest
from openpyxl import load_workbook
from excel_handler import ensure_excel_exists, append_appointment_record

class TestExcelHandler(unittest.TestCase):
    def setUp(self):
        self.test_excel_path = os.path.join(os.path.dirname(__file__), "test_output.xlsx")
        if os.path.exists(self.test_excel_path):
            os.remove(self.test_excel_path)

    def tearDown(self):
        if os.path.exists(self.test_excel_path):
            try:
                os.remove(self.test_excel_path)
            except Exception:
                pass

    def test_excel_creation_and_append(self):
        ensure_excel_exists(self.test_excel_path)
        self.assertTrue(os.path.exists(self.test_excel_path))

        sample_data = {
            "account": "acc_888",
            "password": "pwd_888_secure",
            "time": "2026-07-24 16:00",
            "is_bound": "是",
            "is_diploma": "否"
        }

        append_appointment_record("测试微信名", sample_data, file_path=self.test_excel_path)

        wb = load_workbook(self.test_excel_path)
        ws = wb.active

        # 第 1 行为表头，第 2 行为追加的数据
        self.assertEqual(ws.max_row, 2)

        row_vals = [cell.value for cell in ws[2]]
        self.assertEqual(row_vals[1], "测试微信名")
        self.assertEqual(row_vals[2], "acc_888")
        self.assertEqual(row_vals[3], "pwd_888_secure")
        self.assertEqual(row_vals[4], "2026-07-24 16:00")
        self.assertEqual(row_vals[5], "是")
        self.assertEqual(row_vals[6], "否")

if __name__ == "__main__":
    unittest.main()
